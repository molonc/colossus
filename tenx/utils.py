import os
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from core.utils import GSCForm
from core.constants import chromium_index_mapping
from django.shortcuts import get_object_or_404
from tenx.models import TenxPool

from collections import OrderedDict


def fill_submission_form(form_info):
    """
    Populates template gsc submission with information from form


    Args:
        form_info (dict):           Includes name, email, date and tenx pool id
    
    Returns:
        output_filename (str):      Name of excel file
        workbook (bytes):           Workbook stream

    """

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    gsc_tenx_submission_template = os.path.join(base_dir, "core", "static", "tenx", "gsc_tenx_submission.xlsx")

    form_workbook = load_workbook(gsc_tenx_submission_template, keep_vba=True)
    worksheet = form_workbook["Submission Info"]

    tenxpool = get_object_or_404(TenxPool, pk=form_info["tenxpools"])
    libraries = tenxpool.libraries.all()

    output_filename = f"Aparicio_GSC-0297_Constructed_Library-Submission_Chromium_{tenxpool.pool_name}.xls"

    sublibrary_start_row = 76

    # Fill in submission information: Name, Email, Date
    worksheet.cell(20, 2).value = form_info["name"]
    worksheet.cell(21, 2).value = form_info["email"]
    worksheet.cell(22, 2).value = form_info["date"]


    # Fill in Pool id and tube label in Pool Information
    worksheet.cell(66, 1).value = tenxpool.pool_name
    worksheet.cell(66, 2).value = tenxpool.pool_name

    # Dna Volume is the number of sublibraries times 30
    worksheet.cell(66, 4).value = 30*len(libraries)

    for library in libraries:
        chromium_index = library.tenxlibraryconstructioninformation.index_used.split(",")[0]
        sublibrary_info = {
            'Sub-Library ID': library.name,
            'Tube Label': "N/A",
            'Taxonomy ID': int(library.sample.taxonomy_id),
            'Anonymous Patient ID': library.sample.anonymous_patient_id,
            'Strain': library.sample.strain,
            'Disease Condition/Health Status': library.sample.additionalsampleinformation.disease_condition_health_status,
            'Sex': library.sample.additionalsampleinformation.get_sex_display(),
            'Sample Collection Date': library.tenxlibrarysampledetail.sample_prep_date,
            'Anatomic Site': library.sample.additionalsampleinformation.anatomic_site,
            'Anatomic Sub-Site': library.sample.additionalsampleinformation.anatomic_sub_site,
            'Developmental Stage': library.sample.additionalsampleinformation.developmental_stage,
            'Tissue Type': library.sample.additionalsampleinformation.get_tissue_type_display(),
            'Cell Type (if sorted)': library.sample.additionalsampleinformation.cell_type,
            'Cell Line ID': library.sample.cell_line_id,
            'Pathology/Disease Name (for diseased sample only)': library.sample.additionalsampleinformation.pathology_disease_name,
            'Additional Pathology Information': library.sample.additionalsampleinformation.additional_pathology_info,
            'Grade': library.sample.additionalsampleinformation.grade,
            'Stage': library.sample.additionalsampleinformation.stage,
            'Tumor content (%)': library.sample.additionalsampleinformation.tumour_content,
            'Pathology Occurrence': library.sample.additionalsampleinformation.get_pathology_occurrence_display(),
            'Treatment Status': library.sample.additionalsampleinformation.get_treatment_status_display(),
            'Family Information': library.sample.additionalsampleinformation.family_information,
            'DNA Volume (uL)': "",
            'DNA Concentration (nM)': "",
            'Storage Medium': "",
            'Quantification Method': "",
            'Library Type': library.tenxlibraryconstructioninformation.library_type,
            'Library Construction Method': library.tenxlibraryconstructioninformation.library_construction_method,
            'Size Range (bp)': "",
            'Average Size (bp)': "",
            'Chromium Sample Index Name': chromium_index,
            'Index Read Type (select from drop down list)': "Single Index (i7)",
            'Index Sequence': chromium_index_mapping[chromium_index],
            'No. of cells/IP': None,
            'Crosslinking Method': None,
            'Crosslinking Time': None,
            'Sonication Time': None,
            'Antibody Used': None,
            'Antibody catalogue number': None,
            'Antibody Lot number': None,
            'Antibody Vendor': None,
            'Amount of Antibody Used(ug)': None,
            'Amount of Bead Used(ul)': None,
            'Bead Type': None,
            'Amount of Chromatin Used(ug)': None,
        }

        sublibrary_info = OrderedDict(sublibrary_info)

        for index, key in enumerate(sublibrary_info, start=1):
            worksheet.cell(sublibrary_start_row, index).value = sublibrary_info[key]
        sublibrary_start_row += 1
        
    workbook = save_virtual_workbook(form_workbook)

    return output_filename, workbook


def tenxlibrary_naming_scheme(library):
    if library.well_partition:
        return "_".join(
            ["SCRNA10X", library.chips.lab_name, "CHIP" + str(library.chips.pk).zfill(4), str(library.chip_well).zfill(3)+str(library.well_partition)]
        )
    return "_".join(
        ["SCRNA10X", library.chips.lab_name, "CHIP" + str(library.chips.pk).zfill(4), str(library.chip_well).zfill(3)]
    )


def tenxpool_naming_scheme(pool):
    return "".join(["TENXPOOL", str(pool.id).zfill(4)])
