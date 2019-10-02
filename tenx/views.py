import os
import openpyxl

from django.contrib.auth.decorators import login_required

from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
import requests

from tenx.models import *
from tenx.forms import *
from tenx.utils import tenxpool_naming_scheme, TenXGSCForm
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404, render_to_response, redirect
from django.views.generic import TemplateView

from colossus.settings import LOGIN_URL
from core.views import (LibraryList, LibraryDetail, LibraryDelete, LibraryCreate, LibraryUpdate, SequencingCreate,
                        SequencingDetail, SequencingUpdate, SequencingDelete, LaneCreate, LaneUpdate, LaneDelete)
from core.utils import generate_gsc_form

import json


class TenxLibraryList(LibraryList):
    order = 'sample_id'
    library_class = TenxLibrary
    library_type = 'tenx'


@Render("core/tenxanalysis_list.html")
@login_required
def analys_list(request):
    context = {
        'analyses': TenxAnalysis.objects.all().order_by('id'),
    }
    return context


@Render("core/tenxanalysis_detail.html")
@login_required
def analysis_detail(request, pk):
    analysis = get_object_or_404(TenxAnalysis, pk=pk)
    context = {'analysis': analysis, 'library': analysis.tenx_library, 'sequencings': analysis.tenxsequencing_set}

    tenx_pools = list(map(lambda x: x.tenx_pool, analysis.tenxsequencing_set.all()))
    context['tenx_pools'] = tenx_pools
    return context


@login_required
def library_id_to_pk_redirect(request, pool_id):
    pk = get_object_or_404(TenxLibrary, name=pool_id).pk
    return redirect("/tenx/library/{}".format(pk))


@login_required
def get_gsc_submission_form(request):

    # print(dir(request))
    # print(request.session.keys())
    # print(json.dumps(request.session.decode('utf-8')))
    # print(json.dumps(request.body.decode('utf-8')))

    print("hello")

    if request.method == 'POST':
        form = TenxGSCSubmissionForm(request.POST)

        if form.is_valid():
            print('form is valid')
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            date = form.cleaned_data['date']
            tenxpool_pk = form.cleaned_data['tenxpools']
            print(name, email, date, tenxpool_pk)

            # gsc_form = TenXGSCForm(tenxpool_pk, info=form.cleaned_data)
            # todo: refactor below; ugly
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

            with open(os.path.join(base_dir, "core", "static", "tenx", "gsc_tenx_submission.xlsx"), 'rb') as f:
                filename = "test.xlsx"
                files = {"files": (filename, f)}

                r = requests.post("http://127.0.0.1:8000/tenx/pool/create_form/",
                                  files=files,)

                print(r)


            # create excel here
            # key = f"gsc_form_metadata_{tenxpool_pk}"
            # metadata = request.session.pop(key)
            # ofilename, buffer = generate_gsc_form(tenxpool_pk, metadata)
            # buffer.seek(0)
            # response = HttpResponse(
            #     buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # response['Content-Disposition'] = f'attachment; filename={ofilename}'
            # return response
            return HttpResponseRedirect("/tenx/pool/list")

        else:
            form = TenxGSCSubmissionForm()

    return render(request, 'core/tenx/tenxpool.html', {'form': form})


@csrf_exempt
@login_required
def fill_submission_form(request):

    # TODO: need to define form info

    print("hellooo")
    print(request.FILES)

    return HttpResponse("ok")

    # gsc_tenx_submission_template = os.path.join(
    #     settings.STATIC_URL, "tenx", "gsc_tenx_submission.xlsx")

    # submission_form = workbook.create_sheet("sheetName")
    # form_workbook = openpyxl.load_workbook(gsc_tenx_submission_template)
    # worksheet = form_workbook["Submission Info"]

    # pool_info_row = worksheet[74]
    # pool_start_row = 66

    # sublibrary_info_row = worksheet[74]
    # sublibrary_start_row = 76

    # # Submitting information: Name, Email, Date
    # worksheet.cell(20, 2).value = form_info["name"]
    # worksheet.cell(21, 2).value = form_info["email"]
    # worksheet.cell(22, 2).value = form_info["date"]

    # tenxpool = get_object_or_404(TenxPool, pk=form_info["tenxpools"])
    # libraries = tenxpool.libraries.all()

    # # Dna Volume is the number of sublibraries times 30
    # worksheet.cell(66, 4).value = 30*len(libraries)

    # for library in libraries:
    #     sublibrary_info = {
    #         'Sub-Library ID': library.name,
    #         'Tube Label': "N/A",
    #         'Taxonomy ID': library.sample.taxonomy_id,
    #         'Anonymous Patient ID': library.sample.anonymous_patient_id,
    #         'Strain': library.sample.strain,
    #         'Disease Condition/Health Status': library.sample.additionalsampleinformation.disease_condition_health_status,
    #         'Sex': library.sample.additionalsampleinformation.get_sex_display(),
    #         'Sample Collection Date': library.tenxlibrarysampledetail.sample_prep_date,
    #         'Anatomic Site': library.sample.additionalsampleinformation.anatomic_site,
    #         'Anatomic Sub-Site': library.sample.additionalsampleinformation.anatomic_sub_site,
    #         'Developmental Stage': library.sample.additionalsampleinformation.developmental_stage,
    #         'Tissue Type': library.sample.additionalsampleinformation.get_tissue_type_display(),
    #         'Cell Type (if sorted)': library.sample.additionalsampleinformation.cell_type,
    #         'Cell Line ID': library.sample.cell_line_id,
    #         'Pathology/Disease Name (for diseased sample only)': library.sample.additionalsampleinformation.pathology_disease_name,
    #         'Additional Pathology Information': library.sample.additionalsampleinformation.additional_pathology_info,
    #         'Grade': library.sample.additionalsampleinformation.grade,
    #         'Stage': library.sample.additionalsampleinformation.stage,
    #         'Tumor content (%)': library.sample.additionalsampleinformation.tumour_content,
    #         'Pathology Occurrence': library.sample.additionalsampleinformation.get_pathology_occurrence_display(),
    #         'Treatment Status': library.sample.additionalsampleinformation.get_treatment_status_display(),
    #         'Family Information': library.sample.additionalsampleinformation.family_information,
    #         'DNA Volume (uL)': "",
    #         'DNA Concentration (nM)': "",
    #         'Storage Medium': "",
    #         'Quantification Method': "",
    #         'Library Type': library.tenxlibraryconstructioninformation.library_type,
    #         'Library Construction Method': library.tenxlibraryconstructioninformation.library_construction_method,
    #         'Size Range (bp)': "",
    #         'Average Size (bp)': "",
    #         'Chromium Sample Index Name': library.tenxlibraryconstructioninformation.index_used.split(",")[0],
    #         'Index Read Type (select from drop down list)': "Single Index (i7)",
    #         'Index Sequence': "; ".join(library.tenxlibraryconstructioninformation.index_used.split(",")[1:]),
    #         'No. of cells/IP': None,
    #         'Crosslinking Method': None,
    #         'Crosslinking Time': None,
    #         'Sonication Time': None,
    #         'Antibody Used': None,
    #         'Antibody catalogue number': None,
    #         'Antibody Lot number': None,
    #         'Antibody Vendor': None,
    #         'Amount of Antibody Used(ug)': None,
    #         'Amount of Bead Used(ul)': None,
    #         'Bead Type': None,
    #         'Amount of Chromatin Used(ug)': None,
    #     }

    #     # for cell in worksheet[sublibrary_start_row]:
    #     #     for values in sublibrary_info.values():
    #     #         cell.value =

    #     for values in sublibrary_info.values():
    #         worksheet[sublibrary_start_row] = values
    #         sublibrary_start_row += 1

    # form_workbook.close()


class TenxLibraryDetail(LibraryDetail):
    library_class = TenxLibrary
    library_type = 'tenx'

    def get(self, request, pk):
        library = get_object_or_404(TenxLibrary, pk=pk)
        library_type = 'tenx'

        analyses = TenxAnalysis.objects.filter(tenx_library=pk)

        return self.get_context_and_render(
            request=request,
            library=library,
            library_type=library_type,
            analyses=analyses,
        )


class TenxLibraryDelete(LibraryDelete):
    library_class = TenxLibrary
    library_type = 'tenx'


class TenxLibraryCreate(LibraryCreate):
    lib_form_class = TenxLibraryForm
    libdetail_formset_class = TenxLibrarySampleDetailInlineFormset
    libcons_formset_class = TenxLibraryConstructionInfoInlineFormset
    libqs_formset_class = TenxLibraryQuantificationAndStorageInlineFormset
    library_type = 'tenx'

    def get_context_data(self, pk=None):
        context = super(TenxLibraryCreate, self).get_context_data(pk)
        return context


class TenxLibraryUpdate(LibraryUpdate):
    library_class = TenxLibrary
    lib_form_class = TenxLibraryForm
    libdetail_formset_class = TenxLibrarySampleDetailInlineFormset
    libcons_formset_class = TenxLibraryConstructionInfoInlineFormset
    libqs_formset_class = TenxLibraryQuantificationAndStorageInlineFormset
    library_type = 'tenx'

    def get_context_data(self, pk=None):
        context = super(TenxLibraryUpdate, self).get_context_data(pk)
        library = get_object_or_404(self.library_class, pk=pk)
        context['name'] = library.name

        return context


class TenxPoolList(LoginRequiredMixin, TemplateView):
    login_url = LOGIN_URL
    template_name = "core/tenx/tenxpool_list.html"

    def get_context_data(self):
        context = {
            'pools': TenxPool.objects.all().order_by('id'),
            'form': TenxGSCSubmissionForm(),
        }
        return context


class TenxPoolDetail(TemplateView):

    template_name = "core/tenx/tenxpool_detail.html"

    def get_context_data(self, pk):
        context = {
            'pool': get_object_or_404(TenxPool, pk=pk),
        }
        return context


class TenxPoolDelete(LoginRequiredMixin, TemplateView):
    login_url = LOGIN_URL
    template_name = "core/tenx/tenxpool_delete.html"

    def get_context_data(self, pk):
        context = {
            'pool': get_object_or_404(TenxPool, pk=pk),
            'pk': pk,
        }
        return context

    def post(self, request, pk):
        get_object_or_404(TenxPool, pk=pk).delete()
        msg = "Successfully deleted the Pool."
        messages.success(request, msg)
        return HttpResponseRedirect(reverse('tenx:pool_list'))


class TenxPoolCreate(LoginRequiredMixin, TemplateView):
    login_url = LOGIN_URL
    template_name = "core/tenx/tenxpool_create.html"

    def get_context_data(self):
        context = {'form': TenxPoolForm(), 'tenxlibraries': TenxLibrary.objects.all()}
        return context

    def post(self, request):
        form = TenxPoolForm(request.POST)
        if form.is_valid():
            instance = form.save()
            instance.pool_name = tenxpool_naming_scheme(instance)
            instance.save()
            msg = "Successfully created the %s pool." % instance.pool_name
            messages.success(request, msg)
            return HttpResponseRedirect(instance.get_absolute_url())
        else:
            messages.info(request, form.errors)
            return HttpResponseRedirect(request.get_full_path())


class TenxPoolUpdate(LoginRequiredMixin, TemplateView):
    login_url = LOGIN_URL
    template_name = "core/tenx/tenxpool_update.html"

    def get_context_data(self, pk):
        pool_library = [l.pk for l in get_object_or_404(TenxPool, pk=pk).libraries.all()]
        context = {
            'pk': pk,
            'form': TenxPoolForm(instance=get_object_or_404(TenxPool, pk=pk)),
            'tenxlibraries': TenxLibrary.objects.all(),
            'pool_library': pool_library
        }
        return context

    def post(self, request, pk):
        form = TenxPoolForm(request.POST, instance=get_object_or_404(TenxPool, pk=pk))
        if form.is_valid():
            instance = form.save(commit=False)
            form.save_m2m()
            instance.pool_name = tenxpool_naming_scheme(instance)
            instance.save()

            msg = "Successfully created the %s pool." % instance.pool_name
            messages.success(request, msg)
            return HttpResponseRedirect(instance.get_absolute_url())
        else:
            messages.info(request, form.errors)
            return HttpResponseRedirect(request.get_full_path())


class TenxSequencingList(TemplateView):
    login_url = LOGIN_URL
    template_name = "core/tenx/tenxsequencing_list.html"

    def get_context_data(self):
        context = {
            'sequencings': TenxSequencing.objects.all().order_by('id'),
        }
        return context


class TenxSequencingCreate(SequencingCreate):
    library_class = TenxLibrary
    sequencing_class = TenxSequencing
    form_class = TenxSequencingForm
    library_type = 'tenx'


class TenxSequencingDetail(SequencingDetail):
    sequencing_class = TenxSequencing
    library_type = 'tenx'


class TenxSequencingUpdate(SequencingUpdate):
    sequencing_class = TenxSequencing
    form_class = TenxSequencingForm
    library_type = 'tenx'


class TenxSequencingDelete(SequencingDelete):
    sequencing_class = TenxSequencing
    library_type = 'tenx'


class TenxLaneCreate(LaneCreate):
    sequencing_class = TenxSequencing
    form_class = TenxLaneForm
    library_type = 'tenx'


class TenxLaneUpdate(LaneUpdate):
    lane_class = TenxLane
    form_class = TenxLaneForm
    library_type = 'tenx'


class TenxLaneDelete(LaneDelete):
    lane_class = TenxLane
    library_type = 'tenx'


#============================
# TenxChip views
#----------------------------
class TenxChipCreate(LoginRequiredMixin, TemplateView):

    login_url = LOGIN_URL
    template_name = "core/tenx/tenxchip_create.html"

    def get_context_data(self, **kwargs):
        context = {"form": TenxChipForm}
        return context

    def post(self, request):
        form = TenxChipForm(request.POST)
        if form.is_valid():
            instance = form.save()
            return HttpResponseRedirect(instance.get_absolute_url())
        else:
            return render_to_response('my_template.html', {'form': form})


class TenxChipList(LoginRequiredMixin, TemplateView):

    login_url = LOGIN_URL
    template_name = "core/tenx/tenxchip_list.html"

    def get_context_data(self):
        context = {
            'chips': TenxChip.objects.all().order_by('id'),
        }
        return context


class TenxChipDetail(LoginRequiredMixin, TemplateView):

    login_url = LOGIN_URL
    template_name = "core/tenx/tenxchip_detail.html"

    def get_context_data(self, pk):
        context = {
            'chip': get_object_or_404(TenxChip, pk=pk),
            'pk': pk,
        }

        return context


class TenxChipUpdate(LoginRequiredMixin, TemplateView):

    login_url = LOGIN_URL
    template_name = "core/tenx/tenxchip_update.html"

    def get_context_data(self, pk):
        chip = get_object_or_404(TenxChip, pk=pk)
        form = TenxChipForm(instance=chip)
        context = {"form": form, "pk": pk}
        return context

    def post(self, request, pk):
        chip = get_object_or_404(TenxChip, pk=pk)
        form = TenxChipForm(request.POST, instance=chip)

        if form.is_valid():
            form.save()
            msg = "Successfully updated the Chip."
            messages.success(request, msg)
            return HttpResponseRedirect(chip.get_absolute_url())

        msg = "Failed to update the Chip. Please fix the errors below."
        messages.error(request, msg)
        return self.get_context_and_render(request, form, pk=pk)


class TenxChipDelete(LoginRequiredMixin, TemplateView):

    login_url = LOGIN_URL
    template_name = "core/tenx/tenxchip_delete.html"

    def get_context_data(self, pk):
        context = {
            'chip': get_object_or_404(TenxChip, pk=pk),
        }
        return context

    def post(self, request, pk):
        get_object_or_404(TenxChip, pk=pk).delete()
        msg = "Successfully deleted the Chip."
        messages.success(request, msg)
        return HttpResponseRedirect(reverse('tenx:chip_list'))
